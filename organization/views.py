from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.contenttypes.models import ContentType
from django.db import transaction

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import csv
from django.http import HttpResponse
from django.views.generic import View, ListView

from .models import Unit


# Create your views here.
class OrganizationView(ListView):
    model = Unit
    template_name = 'organization/index.html'
    context_object_name = 'root_units'

    def get_queryset(self):
        # 找出所有頂層單位 (Superior ID 為空)
        return Unit.objects.filter(superior_object_id__isnull=True)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 測試用：在後端印出數量，確認資料庫是否有資料
        print(f"目前資料庫中的單位數量: {self.get_queryset().count()}")
        return context


def download_unit_template(request):
    # 建立活頁簿
    wb = Workbook()
    ws = wb.active
    ws.title = "單位資料匯入範本"

    # 1. 定義表頭 (對應你的模型欄位)
    headers = [
        '單位名稱(name)', '英文名稱(en_name)', '海巡代碼(serial_number)',
        '地址(address)', '電話(landline_phone)', 'Email',
        '上級單位中文名稱(superior_name)', '主管(director)', '主秘(Chief Secretary)'
    ]

    # 寫入表頭並設定樣式
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        # 設定欄寬
        ws.column_dimensions[cell.column_letter].width = 20

    # 2. 寫入一筆範例資料
    sample_data = [
        '北部分署', 'north_branch', '123456',
        '桃園市竹圍路...', '03383xxxx', 'north@cga.gov.tw',
        '', '張長官', '李秘書'
    ]
    ws.append(sample_data)

    # 3. 加入填寫說明
    ws.append([])  # 空行
    ws.append(['填寫說明：'])
    ws.append(['1. 若該單位為最高階層，上級單位名稱請留空。'])
    ws.append(['2. 上級單位必須是已存在於系統中（或在此表上方已定義）的中文名稱。'])

    # 輸出檔案
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="unit_import_template.xlsx"'
    wb.save(response)
    return response


def upload_units(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # 檢查副檔名
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "請上傳 .xlsx 格式的檔案")
            return redirect('organization')

        try:
            result = import_units_from_excel(excel_file)
            messages.success(request, result)
        except Exception as e:
            messages.error(request, f"匯入失敗: {str(e)}")

        return redirect('organization')  # 導回列表頁

    return render(request, 'organization/unit_upload_form.html')


def import_units_from_excel(file_obj):
    """
    使用 en_name 進行上級單位查找的匯入邏輯
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    unit_type = ContentType.objects.get_for_model(Unit)

    # 儲存 {目前單位英名: 上級單位英名}
    hierarchy_map = {}

    with transaction.atomic():
        # 第一階段：建立/更新所有單位
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            # 欄位索引假設：[0]中名, [1]英名, [2]海巡碼, [3]地址, [4]電話, [5]Email, [6]上級英名, [7]主管, [8]主秘
            name = str(row[0]).strip() if row[0] else None
            en_name = str(row[1]).strip() if row[1] else None

            if not name or not en_name:
                continue

                # 1. 建立或更新基本資料
            unit, created = Unit.objects.update_or_create(
                en_name=en_name,
                defaults={
                    'name': name,
                    'serial_number': str(row[2])[:6] if row[2] else None,
                    'address': row[3] if row[3] else "",
                    'landline_phone': row[4],
                    'email': row[5],
                    'director': row[7] if row[7] else "",
                    'deputy_director3': row[8] if row[8] else "",
                }
            )

            # 2. 紀錄上級關係 (此處 row[6] 現在必須填寫上級單位的英名)
            if row[6]:
                hierarchy_map[en_name] = str(row[6]).strip()

        # 第二階段：精準建立上級關聯 (Superior)
        for current_en, superior_en in hierarchy_map.items():
            try:
                # 取得當前單位
                current_unit = Unit.objects.get(en_name=current_en)

                # 直接用英名查找上級，因為 en_name 是唯一值 (Unique)
                superior_obj = Unit.objects.filter(en_name=superior_en).first()

                if superior_obj:
                    current_unit.superior_content_type = unit_type
                    current_unit.superior_object_id = superior_obj.id
                    current_unit.save()
                else:
                    # 可以在這裡加入 log 紀錄找不到的上級英名
                    print(f"警告：找不到上級單位英名 '{superior_en}'")

            except Unit.DoesNotExist:
                continue

    return "匯入完成"
