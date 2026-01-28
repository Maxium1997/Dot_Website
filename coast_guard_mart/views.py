from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.db.models import Sum, Q
from django.utils import timezone
from allauth.socialaccount.models import SocialAccount
from datetime import datetime, timedelta
from collections import defaultdict
from django.urls import reverse
import re, io
import pandas as pd
import qrcode

from .models import Category, Product, ProductAccessory, ProductVariant, \
    WhitelistMember, MemberCredit, CreditTransaction
from Dot_Website.utils import send_line_notification

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from organization.models import Unit


def product_list(request, category_slug=None):
    category = None
    categories = Category.objects.filter(is_public=True)

    products = Product.objects.filter(
        is_display=True,
        category__is_public=True
    ).select_related('category').prefetch_related('variants').order_by('-price')

    if category_slug:
        category = get_object_or_404(categories, slug=category_slug)
        products = products.filter(category=category)

    context = {
        'category': category,
        'categories': categories,
        'products': products,
    }
    return render(request, 'coast_guard_mart/product_list.html', context)


def product_detail(request, pk):
    # 抓取主產品，並預先載入：
    # 1. 主產品的所有規格 (variants)
    # 2. 所有附屬品關聯 (accessory_relations)
    # 3. 附屬品本身的資訊及其規格 (accessory_item__variants)
    # 4. 產品細節圖 (images)
    product = get_object_or_404(
        Product.objects.prefetch_related(
            'variants',
            'images',
            'accessory_relations__accessory_item__variants'
        ),
        pk=pk,
        is_display=True
    )

    # 取得關聯的附屬品資料
    accessories = product.accessory_relations.all()

    context = {
        'product': product,
        'accessories': accessories,
    }
    return render(request, 'coast_guard_mart/product_detail.html', context)


# 獲取使用者當前可用的額度
def get_current_valid_credit(user):
    now = timezone.now()
    return user.credits.filter(
        is_active=True,
        start_date__lte=now,
        end_date__gte=now,
        balance__gt=0
    ).first()   # 取得最新的一張有效卡


# 取得該單位直屬的下級單位
def api_get_subordinates(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    subordinates = unit.get_all_subordinates_direct()
    data = [
        {'id': sub.id, 'name': sub.name} for sub in subordinates
    ]
    return JsonResponse({'results': data})


# 使用者是否可領取點數卡
@login_required
def claim_credit(request):
    current_year = timezone.now().year

    if MemberCredit.objects.filter(user=request.user, fiscal_year=current_year).exists():
        messages.info(request, "您本年度的福利金已領取過。")
        return redirect('coast_guard_mart:product_list')

    top_units = Unit.objects.filter(
        superior_object_id__isnull=True,
        superior_content_type__isnull=True
    ).order_by('name')

    if request.method == 'POST':
        # 按照 level5 -> level4 -> ... -> level1 的順序抓取第一個有值的部分
        selected_unit_id = None
        for i in range(5, 0, -1):
            uid = request.POST.get(f'level{i}')  # 如果您的 HTML level1~4 沒給 name，這行會抓不到
            # 或者簡單點，我們直接強制前端必須完成 level5
            if i == 5:
                uid = request.POST.get('unit')

            if uid and uid.isdigit():
                selected_unit_id = int(uid)
                break

        # 關鍵修正：確保 unit_id 是整數
        raw_unit_id = request.POST.get('unit')
        selected_unit_id = int(raw_unit_id) if raw_unit_id and raw_unit_id.isdigit() else None

        id_number = request.POST.get('id_number', '').strip().upper()
        birthday_str = request.POST.get('birthday')

        # 找出潛在成員
        potential_members = WhitelistMember.objects.filter(
            id_number=id_number,
            birthday=birthday_str,
            is_claimed=False
        )

        print(f"--- Debug: 開始核對 ---")
        print(f"輸入資料: ID={id_number}, 生日={birthday_str}, 選擇單位ID={selected_unit_id}")
        print(f"找到符合身分與生日的白名單數量: {potential_members.count()}")

        target_member = None
        if selected_unit_id and potential_members.exists():
            try:
                selected_unit = Unit.objects.get(id=selected_unit_id)

                for m in potential_members:
                    print(f"檢查成員: {m.name}, 登記單位ID: {m.unit_id}")

                    # 向上追溯：檢查選擇的單位是否在成員登記單位的管轄內
                    curr = selected_unit
                    path = []
                    while curr:
                        path.append(f"{curr.name}({curr.id})")
                        if curr.id == m.unit_id:
                            target_member = m
                            print(f"✅ 匹配成功！路徑匹配到: {curr.name}")
                            break
                        # 向上移動
                        curr = curr.superior

                    print(f"向上追溯路徑: {' -> '.join(path)}")
                    if target_member: break
            except Unit.DoesNotExist:
                print("❌ 錯誤: 找不到選擇的單位")

        if target_member:
            try:
                with transaction.atomic():
                    # 1. 更新白名單狀態
                    target_member.is_claimed = True
                    target_member.claimed_by = request.user
                    target_member.save()

                    # 2. 計算到期日 (假設有效期限到當年度的 12 月 31 日)
                    current_year = timezone.now().year
                    end_of_year = timezone.make_aware(datetime(current_year, 12, 31, 23, 59, 59))

                    # 3. 建立點數卡 (補上 end_date)
                    MemberCredit.objects.create(
                        user=request.user,
                        fiscal_year=current_year,
                        balance=3000,
                        start_date=timezone.now(),
                        end_date=end_of_year,  # <-- 補上這個欄位
                        is_active=True
                    )

                messages.success(request, f"身分核對成功！歡迎 {target_member.name} 同仁，福利金已核發。")
                return redirect('coast_guard_mart:product_list')

            except Exception as e:
                # 這裡會抓到資料庫的限制錯誤
                messages.error(request, f"系統錯誤：{str(e)}")
        else:
            print("❌ 核對最終結果: 失敗")
            messages.error(request, "核對失敗：所選單位與白名單登記資料不符，或該身分已被領取。")

    return render(request, 'coast_guard_mart/claim_credit.html', {'top_units': top_units})


# 加入商品到購物車
def add_to_cart(request, variant_id):
    cart = request.session.get('cart', {})
    quantity = int(request.POST.get('quantity', 1))

    # 將規格 ID 轉為字串作為 key
    v_id = str(variant_id)
    if v_id in cart:
        cart[v_id] += quantity
    else:
        cart[v_id] = quantity

    request.session['cart'] = cart
    messages.success(request, "已加入購物車")
    return redirect('coast_guard_mart:cart_detail')


# 加入多項商品到購物車
def add_to_cart_bulk(request):
    if request.method == 'POST':
        cart = request.session.get('cart', {})

        main_variant_id = request.POST.get('main_variant')
        quantity = int(request.POST.get('quantity', 1))

        # 1. 取得主商品規格物件，用以回推主產品
        main_variant = get_object_or_404(ProductVariant.objects.select_related('product'), id=main_variant_id)
        product = main_variant.product

        # 2. 取得該產品在資料庫中設定的「必選配件數量」
        required_accessory_count = product.accessory_relations.count()

        # 3. 取得使用者提交的配件（過濾掉空值）
        raw_accessory_variants = request.POST.getlist('accessory_variants')
        accessory_ids = sorted([str(aid) for aid in raw_accessory_variants if aid])

        # 4. 核心驗證：如果提交的配件數量不等於要求的數量，代表有漏選
        if len(accessory_ids) < required_accessory_count:
            messages.error(request, "請務必選擇所有加購項目的規格！")
            # 返回原商品頁面
            return redirect('coast_guard_mart:product_detail', pk=product.pk)

        # 5. 建立組合 Key 並存入購物車
        cart_key = "_".join([str(main_variant_id)] + accessory_ids)
        cart[cart_key] = cart.get(cart_key, 0) + quantity

        request.session['cart'] = cart
        messages.success(request, f"【{product.name}】已成功加入購物車")

    return redirect('coast_guard_mart:cart_detail')


# 移除購物車商品
def remove_from_cart(request, cart_key):
    cart = request.session.get('cart', {})
    if cart_key in cart:
        del cart[cart_key]  # 刪除這個 Key，主商品與附屬品會一起消失
        request.session['cart'] = cart
        messages.success(request, "已移除該商品組合")
    return redirect('coast_guard_mart:cart_detail')


# 購物車明細
def cart_detail(request):
    cart = request.session.get('cart', {})
    cart_items = []
    total_price = 0

    for cart_key, qty in cart.items():
        ids = cart_key.split('_')
        try:
            main_variant = ProductVariant.objects.select_related('product').get(id=ids[0])

            # 取得這組裡面的所有附屬品
            accessories = ProductVariant.objects.filter(id__in=ids[1:]).select_related('product')

            # 計算這一組的「單組總額」 (主商品單價 + 所有附屬品單價)
            group_unit_price = main_variant.product.price + sum(acc.product.price for acc in accessories)

            # 計算這一列的總額 (單組總額 * 數量)
            subtotal = group_unit_price * qty
            total_price += subtotal

            cart_items.append({
                'cart_key': cart_key,
                'main_variant': main_variant,
                'accessories': accessories,
                'unit_price': group_unit_price,  # 直接傳入算好的單價
                'quantity': qty,
                'subtotal': subtotal
            })
        except ProductVariant.DoesNotExist:
            continue

    return render(request, 'coast_guard_mart/cart_detail.html', {
        'cart_items': cart_items,
        'total_price': total_price
    })


# 購物車結帳
@login_required
def checkout(request):
    cart = request.session.get('cart', {})
    if not cart:
        messages.warning(request, "您的購物車是空的。")
        return redirect('coast_guard_mart:product_list')

    # 取得當年度有效的福利金卡
    current_year = timezone.now().year
    credit_card = MemberCredit.objects.filter(
        user=request.user,
        fiscal_year=current_year,
        is_active=True
    ).first()

    checkout_items = []
    total_price = 0

    # 1. 組合商品資訊與金額計算 (注意：此處已使用 main_variant.product.price)
    for cart_key, qty in cart.items():
        ids = cart_key.split('_')
        # 預先載入 product 減少資料庫查詢次數
        main_variant = get_object_or_404(ProductVariant.objects.select_related('product'), id=ids[0])
        accessories = ProductVariant.objects.filter(id__in=ids[1:]).select_related('product')

        # 計算單組價格：主商品價格 + 所有配件價格
        group_unit_price = main_variant.product.price + sum(acc.product.price for acc in accessories)
        subtotal = group_unit_price * qty
        total_price += subtotal

        # 建立詳細規格字串 (供後台解析及收據顯示)
        acc_details = [f"{acc.product.name} ({acc.color}/{acc.size})" for acc in accessories]
        spec_info = f"{main_variant.color}/{main_variant.size}"
        if acc_details:
            spec_info += f" [含配件: {' + '.join(acc_details)}]"

        checkout_items.append({
            'name': main_variant.product.name,
            'spec': spec_info,
            'qty': qty,
            'subtotal': subtotal
        })

    # 2. 預先計算預計餘額 (用於前端顯示)
    remaining_balance = 0
    if credit_card:
        remaining_balance = credit_card.balance - total_price

    # 3. 處理 POST 結帳請求
    if request.method == 'POST':
        # 安全檢查：確保卡片存在且餘額足以支付
        if not credit_card or credit_card.balance < total_price:
            messages.error(request, "餘額不足或無效的福利金帳戶。")
            return redirect('coast_guard_mart:cart_detail')

        try:
            with transaction.atomic():
                # 扣除金額
                credit_card.balance -= total_price
                credit_card.save()

                # 組合訂單描述文字內容
                detail_lines = []
                for item in checkout_items:
                    detail_lines.append(f"• {item['name']} - {item['spec']} x {item['qty']}")

                # 產生訂單編號
                order_no = f"CGM{timezone.now().strftime('%Y%m%d%H%M%S')}"

                # 建立交易紀錄
                CreditTransaction.objects.create(
                    credit_card=credit_card,
                    amount=total_price,
                    order_id=order_no,
                    status=CreditTransaction.Status.PREPARING,
                    description="\n".join(detail_lines)
                )

                # --- 新增 LINE 通知 ---
                msg = f"🔔 訂單結帳成功！\n訂單編號：{order_no}\n金額：{total_price} 元\n訂單內容：\n" + "\n".join(
                    detail_lines)
                send_line_notification(request.user, msg)
                # ---------------------

                # 清空購物車
                request.session['cart'] = {}
                request.session.modified = True

                messages.success(request, "結帳成功！")
                return render(request, 'coast_guard_mart/order_success.html', {'order_no': order_no})

        except Exception as e:
            # 發生錯誤時會自動 rollback transaction
            messages.error(request, f"結帳失敗，請聯繫管理員：{str(e)}")

    # 4. 回傳頁面與 Context
    return render(request, 'coast_guard_mart/checkout.html', {
        'checkout_items': checkout_items,
        'total_price': total_price,
        'credit_card': credit_card,
        'remaining_balance': remaining_balance,  # 關鍵：將算好的預計餘額傳給前端
    })


# 訂單列表
@login_required
def order_list(request):
    # 取得該使用者所有的點數卡消費紀錄（從 MemberCredit 關聯過來）
    transactions = CreditTransaction.objects.filter(
        credit_card__user=request.user
    ).order_by('-timestamp')

    return render(request, 'coast_guard_mart/order_list.html', {
        'transactions': transactions
    })


# 訂單明細：本人及管理者均可查閱
@login_required
def order_detail(request, order_id):
    # 先根據 order_id 抓取訂單，不在此時過濾 user
    tx = get_object_or_404(CreditTransaction, order_id=order_id)

    # 權限判定：如果是管理員，或是訂單本人，才允許查看
    if not (request.user.is_staff or tx.credit_card.user == request.user):
        messages.error(request, "您沒有權限查看此訂單。")
        return redirect('coast_guard_mart:product_list')

    return render(request, 'coast_guard_mart/order_detail.html', {'tx': tx})


# 取消訂單：本人及管理者均可取消訂單
@login_required
def cancel_order(request, order_id):
    # 1. 先抓取訂單 (不限本人，管理員也抓得到)
    tx = get_object_or_404(CreditTransaction, order_id=order_id)

    # 2. 權限判斷：必須是本人或是管理員
    is_owner = tx.credit_card.user == request.user
    if not (is_owner or request.user.is_staff):
        messages.error(request, "您沒有權限執行此操作。")
        return redirect('coast_guard_mart:order_list')

    if request.method != 'POST':
        return redirect('coast_guard_mart:order_list')

    # 3. 檢查狀態是否可取消
    if tx.status != CreditTransaction.Status.PREPARING:
        messages.error(request, f"訂單目前的狀態為「{tx.get_status_display()}」，無法取消。")
        return redirect('coast_guard_mart:order_list')

    try:
        with transaction.atomic():
            # 回補金額
            card = tx.credit_card
            card.balance += tx.amount
            card.save()

            # 更新狀態
            tx.status = CreditTransaction.Status.CANCELLED

            # --- 關鍵修正：在訂單描述中加入備註 ---
            if not is_owner:
                # 如果是管理員取消，在原本的商品內容下方追加備註
                admin_note = f"\n\n⚠️ 【系統備註】本訂單已由管理員 {request.user.username} 取消。"
                tx.description += admin_note

            tx.save()

            # --- LINE 通知內容 ---
            msg = (
                f"⚠️ 訂單已取消通知\n"
                f"訂單編號：{tx.order_id}\n"
                f"退還金額：{tx.amount} 元\n"
                f"取消內容：\n{tx.description}"  # 這裡的 description 已包含剛剛加進去的備註
            )

            if not is_owner:
                msg += "\n\n如有相關疑問請洽客服諮詢。"
            else:
                msg += "\n\n點數已退還至您的帳戶。"

            # 發送給訂單主人
            send_line_notification(tx.credit_card.user, msg)
            # ------------------------------------

            messages.success(request, f"訂單 {tx.order_id} 已成功取消。")

    except Exception as e:
        messages.error(request, f"取消操作失敗：{str(e)}")

    if request.user.is_staff:
        return redirect('coast_guard_mart:staff_order_dashboard')
    return redirect('coast_guard_mart:order_list')


def is_staff(user):
    return user.is_staff


@user_passes_test(is_staff)
def staff_order_dashboard(request):
    all_transactions = CreditTransaction.objects.select_related('credit_card__user').order_by('-timestamp')

    # 改用 status 欄位統計，更精準
    total_spent = all_transactions.exclude(status=CreditTransaction.Status.CANCELLED).aggregate(Sum('amount'))['amount__sum'] or 0
    total_orders = all_transactions.count()
    preparing_orders = all_transactions.filter(status=CreditTransaction.Status.PREPARING).count()
    cancelled_orders = all_transactions.filter(status=CreditTransaction.Status.CANCELLED).count()

    return render(request, 'coast_guard_mart/staff/order_dashboard.html', {
        'transactions': all_transactions,
        'total_spent': total_spent,
        'total_orders': total_orders,
        'preparing_orders': preparing_orders,
        'cancelled_orders': cancelled_orders,
    })


@user_passes_test(is_staff)
def staff_inventory_summary(request):
    # 僅統計「備貨中」的訂單
    active_tx = CreditTransaction.objects.filter(status=CreditTransaction.Status.PREPARING)

    main_summary = defaultdict(int)
    acc_summary = defaultdict(int)

    # 正則模式解析
    row_pattern = r"•\s+(.+?)\s+-\s+(.+?)\s+x\s+(\d+)"

    for tx in active_tx:
        rows = re.findall(row_pattern, tx.description)
        for name, full_spec, qty in rows:
            qty = int(qty)

            # 解析配件 [含配件: A (規) + B (規)]
            acc_match = re.search(r"\[含配件:\s*(.+?)\]", full_spec)
            clean_main_spec = full_spec

            if acc_match:
                acc_content = acc_match.group(1)
                for acc_entry in acc_content.split(' + '):
                    if '(' in acc_entry and ')' in acc_entry:
                        # 拆分加購品的 名稱 與 規格
                        a_name, a_spec = acc_entry.rsplit(' (', 1)
                        a_spec = a_spec.replace(')', '')
                        acc_summary[f"{a_name} | {a_spec}"] += qty
                    else:
                        acc_summary[f"{acc_entry} | 無規格"] += qty

                clean_main_spec = re.sub(r"\[含配件:.*?\]", "", full_spec).strip()

            # 統計主商品
            main_summary[f"{name} | {clean_main_spec}"] += qty

    def format_list(d):
        res = []
        for k, v in d.items():
            name, spec = k.split(' | ')
            res.append({'name': name, 'spec': spec, 'total': v})
        return sorted(res, key=lambda x: x['name'])

    return render(request, 'coast_guard_mart/staff/inventory_summary.html', {
        'main_list': format_list(main_summary),
        'acc_list': format_list(acc_summary),
    })


def generate_order_qrcode(request, order_id):
    # 1. 取得絕對路徑
    path = reverse('coast_guard_mart:staff_verify_order_complete', args=[order_id])

    # 2. 手動組合網址，確保使用 https
    # 這樣可以避免 build_absolute_uri 受到代理伺服器錯誤 Header 的影響
    domain = request.get_host()
    verify_url = f"https://{domain}{path}"

    # 偵錯用：如果你在測試環境，可以 print 出來到終端機看網址對不對
    # print(f"QR Code URL: {verify_url}")

    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(verify_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    buffer = BytesIO()
    img.save(buffer, format="PNG")

    return HttpResponse(buffer.getvalue(), content_type="image/png")


@staff_member_required
def staff_verify_order_complete(request, order_id):
    # 1. 抓取訂單
    tx = get_object_or_404(CreditTransaction, order_id=order_id)

    if request.method == 'POST':
        # 2. 檢查是否處於「備貨中」才可核銷
        if tx.status == CreditTransaction.Status.PREPARING:
            try:
                with transaction.atomic():
                    # 變更狀態為已完成
                    tx.status = CreditTransaction.Status.COMPLETED

                    # 3. 在訂單內容追加核銷資訊
                    verify_note = f"\n\n✅ 【系統備註】本訂單已由管理員 {request.user.username} 於 {timezone.now().strftime('%Y-%m-%d %H:%M')} 完成核銷。"
                    tx.description += verify_note
                    tx.save()

                    # 4. 發送 LINE 通知給使用者
                    msg = (
                        f"✅ 訂單核銷完成通知\n"
                        f"訂單編號：{tx.order_id}\n\n"
                        f"您的商品已成功核銷領取！\n"
                        f"核銷人員：{request.user.username}\n"
                        f"核銷時間：{timezone.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                        f"感謝您的購買，祝您使用愉快。"
                    )
                    send_line_notification(tx.credit_card.user, msg)

                messages.success(request, f"訂單 {order_id} 已成功核銷！")
                return redirect('coast_guard_mart:staff_order_dashboard')

            except Exception as e:
                messages.error(request, f"核銷操作失敗：{str(e)}")
        else:
            messages.warning(request, "此訂單狀態已變更，無法重複核銷。")

    return render(request, 'coast_guard_mart/staff/verify_order_complete.html', {'tx': tx})


def clean_spec(spec_string):
    """清理規格字串：移除 /，無顏色時僅顯示尺寸"""
    if not spec_string: return ""
    # 統一將 None 替換為空字串，方便後續判斷
    spec_string = spec_string.replace('None', '').strip()

    if '/' in spec_string:
        parts = spec_string.split('/')
        color = parts[0].strip()
        size = parts[1].strip()
        # 如果顏色部分是空白，只回傳尺寸；否則以空格取代斜線
        if not color or color in ['無', '無顏色']:
            return size
        return f"{color} {size}"
    return spec_string


@user_passes_test(is_staff)
def export_inventory_excel(request):
    # 1. 修正過濾條件：使用模型中定義的 Status.PREPARING
    # 這樣可以確保不論資料庫存的是 'PREPARING' 還是 '備貨中' 都能正確抓取
    active_tx = CreditTransaction.objects.filter(
        status=CreditTransaction.Status.PREPARING
    ).select_related('credit_card__user__whitelist_info__unit')

    # 取得所有顯示中的產品
    display_products = Product.objects.filter(is_active=True, is_display=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "備貨清單"

    # --- 樣式與表頭設定 ---
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    acc_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    base_headers = ["單位", "使用者名稱", "訂單編號", "數量"]
    for i, text in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=i, value=text)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center_align
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        ws.column_dimensions[get_column_letter(i)].width = 25 if i == 1 else 15

    # 動態表頭 (從 E 欄開始)
    header_map = {}
    current_col = 5
    for prod in display_products:
        start_col = current_col
        # 主規格
        ws.cell(row=1, column=current_col, value=prod.name).fill = header_fill
        ws.cell(row=1, column=current_col).font = white_font
        ws.cell(row=1, column=current_col).alignment = center_align
        ws.cell(row=2, column=current_col, value="主規格").alignment = center_align
        header_map[(prod.name, "MAIN")] = current_col
        current_col += 1
        # 配件
        for rel in prod.accessory_relations.all():
            ws.cell(row=1, column=current_col, value=prod.name).fill = header_fill
            ws.cell(row=1, column=current_col).font = white_font
            ws.cell(row=1, column=current_col).alignment = center_align
            cell_acc_sub = ws.cell(row=2, column=current_col, value=rel.accessory_item.name)
            cell_acc_sub.alignment = center_align
            cell_acc_sub.fill = acc_fill
            header_map[(prod.name, rel.accessory_item.name)] = current_col
            current_col += 1
        if current_col - 1 > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=current_col - 1)

    # 2. 填入數據
    row_num = 3
    pattern = r"•\s+(.+?)\s+-\s+(.+?)\s+x\s+(\d+)"

    for tx in active_tx:
        user = tx.credit_card.user
        unit_path = user.whitelist_info.unit.full_path if hasattr(user,
                                                                  'whitelist_info') and user.whitelist_info.unit else ""

        items = re.findall(pattern, tx.description)
        for main_name, full_spec, qty in items:
            ws.cell(row=row_num, column=1, value=unit_path).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_num, column=2, value=user.username).alignment = center_align
            ws.cell(row=row_num, column=3, value=tx.order_id).alignment = center_align
            ws.cell(row=row_num, column=4, value=int(qty)).alignment = center_align

            # 主規格清理與填入
            main_spec_raw = re.sub(r"\[含配件:.*?\]", "", full_spec).strip()
            main_spec_clean = clean_spec(main_spec_raw)
            main_col_idx = header_map.get((main_name, "MAIN"))
            if main_col_idx:
                ws.cell(row=row_num, column=main_col_idx, value=main_spec_clean).alignment = center_align

            # 配件清理與填入
            acc_match = re.search(r"\[含配件:\s*(.+?)\]", full_spec)
            if acc_match:
                acc_entries = acc_match.group(1).split(' + ')
                for entry in acc_entries:
                    try:
                        acc_item_name = entry.split(' (')[0].strip()
                        acc_spec_raw = re.search(r"\((.*?)\)", entry).group(1)
                        acc_spec_clean = clean_spec(acc_spec_raw)
                        acc_col_idx = header_map.get((main_name, acc_item_name))
                        if acc_col_idx:
                            ws.cell(row=row_num, column=acc_col_idx, value=acc_spec_clean).alignment = center_align
                    except:
                        continue
            row_num += 1

    ws.freeze_panes = "E3"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    curr_time = timezone.now().strftime("%Y%m%d_%H%M")
    response['Content-Disposition'] = f'attachment; filename=Preparing_Orders_{curr_time}.xlsx'
    wb.save(response)
    return response


# API: 供前端階層選單調用
@staff_member_required
def api_get_sub_units(request):
    parent_id = request.GET.get('parent_id')

    if parent_id:
        # 取得 Unit 模型的 ContentType
        unit_type = ContentType.objects.get_for_model(Unit)
        # 使用多型欄位進行過濾
        units = Unit.objects.filter(
            superior_object_id=parent_id,
            superior_content_type=unit_type
        ).order_by('name')
    else:
        # 抓取頂層單位
        units = Unit.objects.filter(
            superior_object_id__isnull=True
        ).order_by('name')

    data = [{'id': u.id, 'name': u.name} for u in units]
    return JsonResponse({'results': data})


# 白名單管理主頁
@staff_member_required
def staff_whitelist_manager(request):
    query = request.GET.get('q', '').strip()

    # 使用 select_related 一併抓取 unit 和綁定的使用者資料 (claimed_by)
    members = WhitelistMember.objects.select_related('unit', 'claimed_by') \
        .prefetch_related('unit__superior') \
        .order_by('-id')

    if query:
        members = members.filter(Q(name__icontains=query) | Q(id_number__icontains=query))

    return render(request, 'coast_guard_mart/staff/whitelist_manager.html', {
        'whitelist': members,
        'query': query
    })


# 新增白名單人員
@staff_member_required
def staff_whitelist_add(request):
    if request.method == 'POST':
        unit_id = request.POST.get('unit_id')
        id_num = request.POST.get('id_number', '').strip().upper()
        if not unit_id:
            messages.error(request, "請選擇完整的單位階層。")
        elif WhitelistMember.objects.filter(id_number=id_num).exists():
            messages.error(request, f"身分證 {id_num} 已存在於系統中。")
        else:
            WhitelistMember.objects.create(
                name=request.POST.get('name'),
                id_number=id_num,
                birthday=request.POST.get('birthday'),
                unit_id=unit_id
            )
            messages.success(request, "人員已成功加入白名單。")
            return redirect('coast_guard_mart:staff_whitelist_manager')

    top_units = Unit.objects.filter(
        superior_object_id__isnull=True
    ).order_by('name')
    return render(request, 'coast_guard_mart/staff/whitelist_form.html', {'top_units': top_units})


# 匯出白名單 Excel (含完整組織路徑名稱)
@staff_member_required
def staff_whitelist_export(request):
    members = WhitelistMember.objects.select_related('unit', 'claimed_by').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "白名單清單"

    # 美化表頭
    headers = ["姓名", "身分證字號", "生日", "單位ID", "完整單位路徑", "領取狀態", "綁定帳號"]
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for m in members:
        ws.append([
            m.name,
            m.id_number,
            m.birthday.strftime('%Y-%m-%d') if m.birthday else "",
            m.unit.id if m.unit else "",
            m.unit.full_path if m.unit else "未設定",
            "已領取" if m.is_claimed else "未領取",
            m.claimed_by.username if m.claimed_by else ""
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"CoastGuard_Whitelist_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


# 批次匯入白名單
@staff_member_required
def staff_whitelist_import(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']

        try:
            df = pd.read_excel(file)

            # 更新必要欄位檢查
            required_columns = ['姓名', '身分證字號', '生日', '單位代碼']  # 將單位ID改為代碼
            for col in required_columns:
                if col not in df.columns:
                    raise ValueError(f"Excel 缺少必要欄位：「{col}」")

            df = df.dropna(how='all')
            success_count = 0

            with transaction.atomic():
                for index, row in df.iterrows():
                    # 取得單位代碼並查詢 Unit 物件
                    unit_code = str(row['單位代碼']).strip()
                    try:
                        # 根據 en_name 尋找單位
                        unit_obj = Unit.objects.get(en_name=unit_code)
                    except Unit.DoesNotExist:
                        raise ValueError(f"第 {index + 2} 列錯誤：找不到單位代碼為「{unit_code}」的單位。")

                    id_num = str(row['身分證字號']).strip().upper()
                    birthday = pd.to_datetime(row['生日']).date()

                    WhitelistMember.objects.update_or_create(
                        id_number=id_num,
                        defaults={
                            'unit': unit_obj,  # 直接傳入物件
                            'name': str(row['姓名']).strip(),
                            'birthday': birthday,
                        }
                    )
                    success_count += 1

            messages.success(request, f"匯入成功！共新增/更新 {success_count} 筆資料。")
            return redirect('coast_guard_mart:staff_whitelist_manager')

        except Exception as e:
            messages.error(request, f"匯入失敗：{str(e)}")

    return render(request, 'coast_guard_mart/staff/whitelist_import.html')


# 提供批次匯入 Excel 範例檔
@staff_member_required
def download_whitelist_template(request):
    # 1. 範例資料 (使用 en_name 作為代碼)
    data = [
        {'姓名': '王小明', '身分證字號': 'A123456789', '生日': '1990-01-01', '單位代碼': 'CGA_HQS'},
        {'姓名': '李美華', '身分證字號': 'B223456789', '生日': '1985-05-20', '單位代碼': 'S_BRANCH'}
    ]
    df_sample = pd.DataFrame(data)

    # 2. 製作單位對照表 (從資料庫抓取所有 Unit)
    units = Unit.objects.all().order_by('en_name')
    unit_lookup_data = [
        {'單位中文名稱': u.full_path, '單位代碼 (en_name)': u.en_name}
        for u in units if u.en_name  # 只列出有設定英文名稱的單位
    ]
    df_units = pd.DataFrame(unit_lookup_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_sample.to_excel(writer, index=False, sheet_name='匯入名單填寫')
        df_units.to_excel(writer, index=False, sheet_name='單位代碼對照表')

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="whitelist_import_sample.xlsx"'
    return response


# 刪除白名單人員
@staff_member_required
def staff_whitelist_delete(request):
    if request.method == 'POST':
        member = get_object_or_404(WhitelistMember, id=request.POST.get('member_id'))
        member.delete()
        messages.success(request, f"人員 {member.name} 已從白名單移除。")
    return redirect('coast_guard_mart:staff_whitelist_manager')
